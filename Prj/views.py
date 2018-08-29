from django.shortcuts import render, redirect
from django.contrib import auth
from django.urls import reverse
from .forms import LoginForm, RegForm
from django.contrib.auth.models import User
from django.http import HttpResponseRedirect, HttpResponse, JsonResponse
from openpyxl import Workbook #写
import io
from mainform.models import Questionnaire, Question, Answer
from urllib import parse
# from .settings import BASE_DIR
from .forms import UploadFileForm
from openpyxl import load_workbook #读

def home(request):
    context = {}
    return render(request, 'home.html', context)

# Create your views here.
def login(request):
    '''username = request.POST.get('username', '')
    password = request.POST.get('password', '')
    user = auth.authenticate(request, username=username, password=password)
    referer = request.META.get('HTTP_REFERER', reverse('home'))

    if user is not None:
        auth.login(request, user)
        return redirect(referer)
    else:
        return render(request, 'error.html', {'message': '用户名或密码不正确'})'''
    if request.method == 'POST':
        login_form = LoginForm(request.POST)
        if login_form.is_valid():
                user = login_form.cleaned_data['user']
                auth.login(request, user)
                return redirect(request.GET.get('from', reverse('home')))
    else:
        login_form = LoginForm()
    context = {}
    context['login_form'] = login_form
    return render(request, 'login.html', context)

def register(request):
    if request.method == 'POST':
        reg_form = RegForm(request.POST)
        if reg_form.is_valid():
            username = reg_form.cleaned_data['username']
            email = reg_form.cleaned_data['email']
            password = reg_form.cleaned_data['password']
            # 创建用户
            '''user = User.objects.create_user(username, email, password)
            user.save()'''
            # 登录用户
            user = User()
            user.username = username
            user.email = email
            user.set_password(password)
            user.save()

            user = auth.authenticate(username=username, password=password)
            auth.login(request, user)
            return redirect(request.GET.get('from', reverse('home')))
    else:
        reg_form = RegForm()
    context = {}
    context['reg_form'] = reg_form
    return render(request, 'register.html', context)

def logout(request):
    auth.logout(request)
    return HttpResponseRedirect("/login/")

def statistics(request):
    if request.user.is_authenticated:
        if not request.user.username == 'yb':
            return render(request, 'message.html', {'message': '你不是管理员！'})
        else:
            context = {}
            question_sheet_all = Questionnaire.objects.all()
            context['list'] = question_sheet_all
            return render(request, 'static-list.html', context)
    else:
        return render(request, 'message.html', {'message': '请先登录'})

def statistics_detail(request, static_pk):
    workbook = Workbook()
    sheet1 = workbook.active
    question_paper = Questionnaire.objects.get(pk=static_pk)
    sheet1.title = question_paper.paper_name
    questions = Question.objects.filter(question_sheet=question_paper)
    answers = Answer.objects.filter(question_sheet=question_paper)
    authors = User.objects.all()
    row_index = 1
    col_index = 1
    for question in questions:
        sheet1.cell(row=row_index, column=col_index).value = question.question_name
        answers_to = answers.filter(question=question)
        for author in authors:
            row_index = row_index + 1
            answer = answers_to.filter(author=author)[0]
            sheet1.cell(row=row_index, column=col_index).value = answer.answer_name
        col_index = col_index + 1
        row_index = 1
        # 写入IO
    res = get_excel_stream(workbook)
    # 设置HttpResponse的类型
    response = HttpResponse(content_type='application/vnd.ms-excel')
    response['Content-Disposition'] = 'attachment;filename=' + parse.quote(question_paper.paper_name) + '.xls'
    # 将文件流写入到response返回
    response.write(res)
    return response

def get_excel_stream(file):
    # StringIO操作的只能是str，如果要操作二进制数据，就需要使用BytesIO。
    excel_stream = io.BytesIO()
    # 这点很重要，传给save函数的不是保存文件名，而是一个BytesIO流（在内存中读写）
    file.save(excel_stream)
    # getvalue方法用于获得写入后的byte将结果返回给re
    res = excel_stream.getvalue()
    excel_stream.close()
    return res


def upload(request):
    if request.method == 'POST':
        upload_form = UploadFileForm(request.POST, request.FILES)
        if upload_form.is_valid():
            title = upload_form.cleaned_data['title']
            form = upload_form.cleaned_data['file']
            is_row = upload_form.cleaned_data['is_row']

            #表格处理
            workbook = load_workbook(form)
            sheet1 = workbook.active  # 正在活动的sheet

            # data = {}
            # i = 0

            questions = []
            # 遍历得到每行第一个单元格的值
            if not is_row:
                for row in sheet1.rows:
                    cell = row[0].value
                    questions.append(cell)
            else:
                for col in sheet1.columns:
                    cell = col[0].value
                    questions.append(cell)
                # data[str(i)] = cell
                # i = i + 1
            if not questions[0] == '':
                question_paper = Questionnaire()
                question_paper.paper_name = title
                question_paper.author = request.user
                question_paper.save()
                for question in questions:
                    ques = Question()
                    ques.question_name = question
                    ques.author = request.user
                    ques.question_sheet = question_paper
                    ques.save()
                return render(request, 'message.html', {'message': '创建成功！'})
            else:
                return render(request, 'message.html', {'message': '创建失败！'})
            # print(data)
            # return JsonResponse(data)
            # handle_uploaded_file(form)
            # return render(request, 'message.html', {'message':'读取成功！'})
    else:
        upload_form = UploadFileForm()
    context = {}
    context['upload_form'] = upload_form
    return render(request, 'upload.html', context)
